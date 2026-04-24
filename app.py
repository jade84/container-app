# --- IMPORT ---
from flask import Flask, render_template, request, redirect
from models import get_connection
from datetime import datetime, date
import pandas as pd
from flask import send_file
import json
from db import query_one, query_all, execute
from dotenv import load_dotenv
load_dotenv()
from services.container_service import get_container_list, get_shipping_lines
from services.container_service import find_container, export_container
from services.container_service import create_container

app = Flask(__name__)
@app.template_filter('from_json')
def from_json(value):
    return json.loads(value)
# --- HÀM CHECK ISO ---
def is_valid_container(container_no):
    import re

    container_no = container_no.upper()

    if not re.match(r'^[A-Z]{4}\d{7}$', container_no):
        return False

    values = {
        'A':10,'B':12,'C':13,'D':14,'E':15,'F':16,'G':17,'H':18,
        'I':19,'J':20,'K':21,'L':23,'M':24,'N':25,'O':26,'P':27,
        'Q':28,'R':29,'S':30,'T':31,'U':32,'V':34,'W':35,'X':36,
        'Y':37,'Z':38
    }

    total = 0

    for i, char in enumerate(container_no[:10]):
        if char.isalpha():
            value = values[char]
        else:
            value = int(char)

        total += value * (2 ** i)

    check_digit = total % 11
    if check_digit == 10:
        check_digit = 0

    return check_digit == int(container_no[-1])
# --- ROUTE: CONTAINER in ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/in', methods=['GET', 'POST'])
def container_in():
    from datetime import date

    lines = query_all("SELECT * FROM shipping_lines ORDER BY code")

    if request.method == 'POST':
        result = create_container(
            request.form['container_no'],
            request.form['shipping_line'],
            request.form['size'],
            request.form['status'],
            request.form['date_in']
        )

        if 'error' in result:
            return render_template(
                'in.html',
                error=result['error'],
                form=request.form,
                today=date.today().strftime('%Y-%m-%d'),
                lines=lines
            )

        return redirect('/list')

    return render_template(
        'in.html',
        form={},
        today=date.today().strftime('%Y-%m-%d'),
        lines=lines
    )
# --- ROUTE: API ---

@app.route('/check_container')
def check_container():
    container_no = request.args.get('container_no')

    result = query_one("""
        SELECT id FROM containers
        WHERE container_no=%s AND date_out IS NULL
    """, (container_no,))

    return {'exists': result is not None}

# --- ROUTE: CONTAINER List ---

from services.container_service import get_container_list, get_shipping_lines
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/list')
def container_list():
    query = request.args.get('q')
    line = request.args.get('line')

    data = get_container_list(query, line)
    lines = get_shipping_lines()

    return render_template(
        'list.html',
        title="Tồn bãi",
        date_column='Date In',
        containers=data,
        q=query,
        line=line,
        lines=lines
    )
# --- ROUTE: CONTAINER OUT ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')


@app.route('/out', methods=['GET', 'POST'])
def container_out():

    data = None

    if request.method == 'POST':
        container_no = request.form['container_no'].upper()

        # 👉 nếu có booking → là bước xuất
        if 'booking_no' in request.form:
            booking_no = request.form['booking_no']
            customer_name = request.form['customer_name']

            result = export_container(container_no, booking_no, customer_name)

            if 'error' in result:
                return render_template('out.html', error=result['error'])

            return render_template('out.html', success=result['success'])

        # 👉 bước tìm
        data = find_container(container_no)

    return render_template('out.html', data=data)
# --- ROUTE:  history ---
from services.container_service import get_container_history
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/history')
def container_history():
    data = get_container_history()

    return render_template(
        'list.html',
        containers=data,
        title="Xuất bãi",
        date_column='Date Out',
        q=None,
        line=None,
        lines=[]
    )
# --- ROUTE:  export ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/export')
def export_excel():
    import io
    import pandas as pd
    from datetime import date, datetime
    from flask import send_file

    query = request.args.get('q') or ''
    line = request.args.get('line') or ''

    sql = """
    SELECT 
        container_no,
        shipping_line,
        size,
        status,
        date_in,
        date_out,
        booking_no,
        customer_name
    FROM containers
    WHERE date_out IS NULL
    """
    params = []

    if query:
        sql += " AND container_no LIKE %s"
        params.append('%' + query.upper() + '%')

    if line:
        sql += " AND shipping_line LIKE %s"
        params.append('%' + line + '%')

    # ✅ dùng query_all (đã refactor)
    data = query_all(sql, params)

    # --- DataFrame ---
    df = pd.DataFrame(data)

    # ❗ nếu không có data → tránh lỗi
    if df.empty:
        df = pd.DataFrame(columns=[
            'container_no','shipping_line','size','status',
            'date_in','date_out','booking_no','customer_name'
        ])

    # --- Days ---
    df['Days'] = df['date_in'].apply(
        lambda x: (date.today() - x).days if x else 0
    )
    df['Days'] = df['Days'].apply(lambda x: x if x > 0 else 1)

    # --- Rename ---
    df = df.rename(columns={
        'container_no': 'Container No',
        'shipping_line': 'Shipping Line',
        'size': 'Size',
        'status': 'Status',
        'date_in': 'Date In',
        'date_out': 'Date Out',
        'booking_no': 'Booking',
        'customer_name': 'Customer'
    })

    # --- STT ---
    df.insert(0, 'STT', range(1, len(df) + 1))

    # ========================
    # 👉 EXCEL MEMORY
    # ========================
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    # --- ghi data ---
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # --- căn giữa ---
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- header ---
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)

    ws.row_dimensions[1].height = 30

    # --- auto width ---
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # --- highlight Days ---
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    days_col = None
    for col in ws[1]:
        if col.value == 'Days':
            days_col = col.column
            break

    if days_col:
        for row in ws.iter_rows(min_row=2):
            cell = row[days_col - 1]

            try:
                days = int(cell.value)
            except:
                continue

            if days >= 15:
                cell.fill = red_fill
            elif days >= 7:
                cell.fill = orange_fill

    # --- freeze ---
    ws.freeze_panes = "B2"

    # --- table ---
    end_row = ws.max_row
    end_col = ws.max_column
    table_range = f"A1:{ws.cell(row=end_row, column=end_col).coordinate}"

    table = Table(displayName="ContainerTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    # --- save ---
    wb.save(output)
    output.seek(0)

    filename = f"containers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# --- ROUTE:  edit ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_container(id):
    import json

    # --- GET container ---
    container = query_one("SELECT * FROM containers WHERE id=%s", (id,))

    if not container:
        return "Không tìm thấy container!"

    if request.method == 'POST':
        shipping_line = request.form['shipping_line']
        size = request.form['size']
        status = request.form['status']

        # --- dữ liệu cũ ---
        old = container

        # --- UPDATE ---
        execute("""
            UPDATE containers
            SET shipping_line=%s,
                size=%s,
                status=%s
            WHERE id=%s
        """, (shipping_line, size, status, id))

        # --- LOG ---
        changes = {}

        fields = {
            'shipping_line': 'Shipping Line',
            'size': 'Size',
            'status': 'Status'
        }

        for field, label in fields.items():
            old_val = old[field]
            new_val = locals()[field]

            if old_val != new_val:
                changes[label] = f"{old_val} → {new_val}"

        if changes:
            execute("""
                INSERT INTO container_logs (container_id, action, old_data, new_data)
                VALUES (%s, %s, %s, %s)
            """, (
                id,
                "UPDATE",
                json.dumps(old, default=str, ensure_ascii=False),
                json.dumps(changes, ensure_ascii=False)
            ))

        return redirect('/list')

    # --- GET ---
    return render_template('edit.html', c=container)
# --- ROUTE:  logs ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/logs')
def view_logs():
    logs = query_all("""
        SELECT 
            l.*, c.container_no
        FROM container_logs l
        LEFT JOIN containers c ON l.container_id = c.id
        ORDER BY l.created_at DESC
    """)

    return render_template('logs.html', logs=logs)
# --- ROUTE:  dashboad ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/dashboard')
def dashboard():
    from datetime import date

    # ✅ Tổng tồn
    total_row = query_one("""
        SELECT COUNT(*) as total
        FROM containers
        WHERE date_out IS NULL
    """)
    total = total_row['total'] if total_row else 0

    # ✅ Theo hãng tàu
    by_line = query_all("""
        SELECT shipping_line, COUNT(*) as count
        FROM containers
        WHERE date_out IS NULL
        GROUP BY shipping_line
        ORDER BY count DESC
    """)

    # ✅ Lấy date_in để tính ngày tồn
    rows = query_all("""
        SELECT date_in
        FROM containers
        WHERE date_out IS NULL
    """)

    over_7 = 0
    over_15 = 0

    today = date.today()

    for r in rows:
        if r['date_in']:
            days = (today - r['date_in']).days

            if days >= 15:
                over_15 += 1
            elif days >= 7:
                over_7 += 1

    return render_template(
        'dashboard.html',
        total=total,
        by_line=by_line,
        over_7=over_7,
        over_15=over_15
    )
# --- ROUTE:  Hãng tàu/ add ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/shipping_lines', methods=['GET', 'POST'])
def shipping_lines():
    if request.method == 'POST':
        code = request.form['code'].upper().strip()
        name = request.form['name'].strip()

        # ❗ CHECK RỖNG
        if not code or not name:
            lines = query_all("SELECT * FROM shipping_lines ORDER BY code")
            return render_template(
                'shipping_lines.html',
                lines=lines,
                error="Thiếu thông tin!"
            )

        # ❗ CHECK TRÙNG
        existing = query_one(
            "SELECT code FROM shipping_lines WHERE code=%s",
            (code,)
        )

        if existing:
            lines = query_all("SELECT * FROM shipping_lines ORDER BY code")
            return render_template(
                'shipping_lines.html',
                lines=lines,
                error="Code đã tồn tại!"
            )

        # ✅ INSERT
        execute(
            "INSERT INTO shipping_lines (code, name) VALUES (%s, %s)",
            (code, name)
        )

        return redirect('/shipping_lines')

    # --- GET ---
    lines = query_all("SELECT * FROM shipping_lines ORDER BY code")

    return render_template('shipping_lines.html', lines=lines)
# --- ROUTE:  Hãng tàu/edit ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/shipping_lines/edit/<code>', methods=['GET', 'POST'])
def edit_shipping_line(code):

    # --- GET dữ liệu ---
    line = query_one(
        "SELECT * FROM shipping_lines WHERE code=%s",
        (code,)
    )

    if not line:
        return "Không tìm thấy hãng tàu!"

    if request.method == 'POST':
        name = request.form['name'].strip()

        # ❗ CHECK RỖNG
        if not name:
            return render_template(
                'edit_shipping_line.html',
                line=line,
                error="Tên không được để trống!"
            )

        # ✅ UPDATE
        execute(
            "UPDATE shipping_lines SET name=%s WHERE code=%s",
            (name, code)
        )

        return redirect('/shipping_lines')

    return render_template('edit_shipping_line.html', line=line)
# --- ROUTE:  Hãng tàu/delete ---
from flask import redirect

@app.route('/')
def home():
    return redirect('/list')

@app.route('/shipping_lines/delete/<code>')
def delete_shipping_line(code):

    # ❗ CHECK đang được dùng
    used = query_one("""
        SELECT 1 FROM containers
        WHERE shipping_line=%s LIMIT 1
    """, (code,))

    if used:
        return "Không thể xóa: đang có container dùng code này!"

    # ✅ DELETE
    execute(
        "DELETE FROM shipping_lines WHERE code=%s",
        (code,)
    )

    return redirect('/shipping_lines')
# --- RUN APP ---
import os

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)